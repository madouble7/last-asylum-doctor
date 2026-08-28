"""Strict reader for the finalized Shop Doctor workbook format."""
# ruff: noqa: E501

from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from last_asylum_doctor.models.economic import (
    CurrencyAssumption,
    ShopDoctorWorkbook,
    WorkbookCashPack,
    WorkbookItem,
    WorkbookOffer,
    WorkbookPackComponent,
    WorkbookRelationship,
)


class ShopDoctorWorkbookError(ValueError):
    """Raised when a workbook cannot be interpreted without guessing."""


_SHEETS = {
    "LIVE OFFERS",
    "CASH PACKS",
    "PACK CONTENTS",
    "ITEM CATALOG",
    "CALC ENGINE",
    "ADVANCED SETTINGS",
}
_ALIAS_SPLIT = re.compile(r"\s*(?:,|;|\n)\s*")


def inspect_shop_doctor_workbook(path: Path) -> ShopDoctorWorkbook:
    """Read and validate raw workbook facts without changing the source file."""
    if not path.is_file():
        raise ShopDoctorWorkbookError(f"Workbook does not exist: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ShopDoctorWorkbookError("Shop Doctor source must be an .xlsx workbook")

    with path.open("rb") as source_file:
        checksum = hashlib.file_digest(source_file, "sha256").hexdigest()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        missing = sorted(_SHEETS.difference(workbook.sheetnames))
        if missing:
            raise ShopDoctorWorkbookError(
                f"Workbook is missing required sheets: {', '.join(missing)}"
            )
        item_rows = tuple(_rows(workbook["ITEM CATALOG"]))
        items = tuple(_item_from_row(row, values) for row, values in item_rows)
        _assert_unique([item.name for item in items], "canonical item names")
        item_names = {item.name for item in items}
        offers = tuple(
            _offer_from_row(row, values)
            for row, values in _rows(workbook["LIVE OFFERS"])
        )
        packs = tuple(
            _pack_from_row(row, values) for row, values in _rows(workbook["CASH PACKS"])
        )
        components = tuple(
            _component_from_row(row, values)
            for row, values in _rows(workbook["PACK CONTENTS"])
        )
        assumptions = tuple(
            _currency_from_row(row, values)
            for row, values in _advanced_currency_rows(workbook["ADVANCED SETTINGS"])
        )
        relationships = tuple(_relationships_from_item(item_rows))
        _validate(items, offers, packs, components, relationships, item_names)
        model_observations = tuple(_model_observations(workbook["CALC ENGINE"]))
        snapshot_date = max(
            [offer.date_seen for offer in offers] + [pack.date_seen for pack in packs]
        )
        return ShopDoctorWorkbook(
            filename=path.name,
            sha256=checksum,
            size_bytes=path.stat().st_size,
            sheet_names=tuple(workbook.sheetnames),
            snapshot_date=snapshot_date,
            items=items,
            offers=offers,
            cash_packs=packs,
            pack_components=components,
            relationships=relationships,
            currency_assumptions=assumptions,
            model_observations=model_observations,
        )
    finally:
        workbook.close()


def _rows(sheet: Any) -> Iterable[tuple[int, dict[str, Any]]]:
    headers = [
        str(value or "").strip()
        for value in next(sheet.iter_rows(min_row=5, max_row=5, values_only=True))
    ]
    if not any(headers):
        raise ShopDoctorWorkbookError(f"{sheet.title}: header row 5 is empty")
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=6, values_only=True), start=6
    ):
        if not row or row[0] in (None, ""):
            continue
        yield (
            row_number,
            {
                headers[index]: value
                for index, value in enumerate(row)
                if index < len(headers)
            },
        )


def _value(values: dict[str, Any], key: str, *, required: bool = False) -> Any:
    value = values.get(key)
    if required and (value is None or (isinstance(value, str) and not value.strip())):
        raise ShopDoctorWorkbookError(f"Required workbook value is blank: {key}")
    return value


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _number(value: Any, *, field: str) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ShopDoctorWorkbookError(f"{field} must be numeric")
    try:
        return float(value)
    except (TypeError, ValueError) as error:
        raise ShopDoctorWorkbookError(f"{field} must be numeric") from error


def _date(value: Any, *, field: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise ShopDoctorWorkbookError(f"{field} must be an Excel date")


def _item_from_row(row: int, values: dict[str, Any]) -> WorkbookItem:
    aliases = tuple(
        alias
        for alias in _ALIAS_SPLIT.split(
            _text(values.get("Aliases / Equivalencies")) or ""
        )
        if alias
    )
    model_values = {
        "strategic_tier": _text(values.get("Strategic Tier")),
        "personal_priority": _text(values.get("Personal Priority")),
        "intrinsic_progression": _number(
            values.get("Intrinsic Progression"), field="Intrinsic Progression"
        ),
        "effective_ad_points": _number(
            values.get("Effective AD Pts / Base"), field="Effective AD Pts / Base"
        ),
        "published_ad_points": _number(
            values.get("Published Base AD Pts / Base"),
            field="Published Base AD Pts / Base",
        ),
        "verification": _text(values.get("Verification")),
    }
    return WorkbookItem(
        name=str(_value(values, "Canonical Item", required=True)).strip(),
        base_unit=str(_value(values, "Base Unit", required=True)).strip(),
        category=_text(values.get("Category")),
        item_type=_text(values.get("Item Type")),
        data_status=_text(values.get("Data Status")),
        source=_text(values.get("Source")),
        notes=_text(values.get("Notes")),
        aliases=aliases,
        doctor_resource_key=_text(values.get("Doctor Resource Key")),
        source_row=row,
        model_values=model_values,
    )


def _offer_from_row(row: int, values: dict[str, Any]) -> WorkbookOffer:
    return WorkbookOffer(
        offer_id=str(_value(values, "Offer ID", required=True)).strip(),
        active=_text(values.get("Active? (optional)")),
        date_seen=_date(_value(values, "Date Seen", required=True), field="Date Seen"),
        shop_name=str(_value(values, "Shop / Event", required=True)).strip(),
        shop_snapshot_id=str(_value(values, "Shop Snapshot ID", required=True)).strip(),
        item_name=str(_value(values, "Canonical Item", required=True)).strip(),
        normalized_quantity=_required_number(values, "Normalized Qty (Base Units)"),
        cost=_required_number(values, "Cost"),
        currency_name=str(_value(values, "Currency", required=True)).strip(),
        direct_cash_price=_number(values.get("Direct Cash $"), field="Direct Cash $"),
        repeat_limit=_number(values.get("Repeat Limit"), field="Repeat Limit"),
        notes=_text(values.get("Notes")),
        source=_text(values.get("Source")),
        actual_package_item=_text(values.get("Actual Package Item")),
        actual_item_count=_number(
            values.get("Actual Item Count"), field="Actual Item Count"
        ),
        base_units_per_item=_number(
            values.get("Base Units per Item"), field="Base Units per Item"
        ),
        normalized_quantity_check=_number(
            values.get("Normalized Qty Check"), field="Normalized Qty Check"
        ),
        speedup_type=_text(values.get("Speedup Type")),
        package_display=_package_display(
            _text(values.get("Actual Package Item")),
            _number(values.get("Actual Item Count"), field="Actual Item Count"),
        ),
        source_row=row,
    )


def _pack_from_row(row: int, values: dict[str, Any]) -> WorkbookCashPack:
    bonus = _number(
        values.get("Assumed Bonus Diamonds"), field="Assumed Bonus Diamonds"
    )
    return WorkbookCashPack(
        pack_id=str(_value(values, "Pack ID", required=True)).strip(),
        active=_text(values.get("Active? (optional)")),
        date_seen=_date(_value(values, "Date Seen", required=True), field="Date Seen"),
        name=str(_value(values, "Pack", required=True)).strip(),
        price_usd=_number(values.get("Price $"), field="Price $"),
        assumed_bonus_diamonds=int(bonus) if bonus is not None else None,
        snapshot_id=str(_value(values, "Snapshot ID", required=True)).strip(),
        repeat_limit=_number(values.get("Repeat Limit"), field="Repeat Limit"),
        notes=_text(values.get("Notes")),
        source=_text(values.get("Source")),
        valuation_status=_text(values.get("Valuation Status")),
        components_valued=_text(values.get("Components Valued")),
        source_row=row,
    )


def _component_from_row(row: int, values: dict[str, Any]) -> WorkbookPackComponent:
    return WorkbookPackComponent(
        pack_id=str(_value(values, "Pack ID", required=True)).strip(),
        item_name=str(_value(values, "Canonical Item", required=True)).strip(),
        normalized_quantity=_required_number(values, "Normalized Qty (Base Units)"),
        notes=_text(values.get("Notes")),
        actual_package_item=_text(values.get("Actual Package Item")),
        actual_item_count=_number(
            values.get("Actual Item Count"), field="Actual Item Count"
        ),
        base_units_per_item=_number(
            values.get("Base Units per Item"), field="Base Units per Item"
        ),
        normalized_quantity_check=_number(
            values.get("Normalized Qty Check"), field="Normalized Qty Check"
        ),
        speedup_type=_text(values.get("Speedup Type")),
        data_status=_text(values.get("Data Status")),
        package_display=_text(values.get("Package Display")),
        normalization_check=_text(values.get("Normalization Check")),
        source_row=row,
    )


def _advanced_currency_rows(sheet: Any) -> Iterable[tuple[int, dict[str, Any]]]:
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=6, max_col=10, values_only=True), start=6
    ):
        if len(row) < 9 or row[4] in (None, ""):
            continue
        yield (
            row_number,
            {
                "Currency": row[4],
                "VE$ / Unit": row[5],
                "Confidence": row[6],
                "Basis": row[7],
                "Active?": row[8],
                "Notes": row[9] if len(row) > 9 else None,
            },
        )


def _currency_from_row(row: int, values: dict[str, Any]) -> CurrencyAssumption:
    return CurrencyAssumption(
        currency_name=str(_value(values, "Currency", required=True)).strip(),
        value_per_unit=_number(values.get("VE$ / Unit"), field="VE$ / Unit"),
        classification=str(_value(values, "Confidence", required=True)).strip(),
        basis=_text(values.get("Basis")),
        active=_text(values.get("Active?")),
        notes=_text(values.get("Notes")),
        source_row=row,
    )


def _relationships_from_item(
    item_rows: Iterable[tuple[int, dict[str, Any]]],
) -> Iterable[WorkbookRelationship]:
    for row, values in item_rows:
        name = _text(values.get("Canonical Item"))
        rule = _text(values.get("Contents Rule"))
        if not name or not rule:
            continue
        for number in range(1, 5):
            item = _text(values.get(f"Choice {number} Item"))
            quantity = _number(
                values.get(f"Choice {number} Qty (Base Units)"), field="Choice quantity"
            )
            if item is not None and quantity is not None:
                yield WorkbookRelationship(
                    container_name=name,
                    context=_text(values.get("Yield Basis")),
                    contents_rule=rule,
                    option_name=item,
                    option_quantity=quantity,
                    source_row=row,
                )


def _model_observations(
    sheet: Any,
) -> Iterable[tuple[str, str, float | str | None, int]]:
    headers = list(next(sheet.iter_rows(min_row=5, max_row=5, values_only=True)))
    positions = {
        str(value): index for index, value in enumerate(headers) if value is not None
    }
    metrics = (
        "VE$ / Package",
        "VE$ / Unit",
        "Historical Best Price",
        "Historical Median",
    )
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=6, values_only=True), start=6
    ):
        if not row or row[0] in (None, ""):
            continue
        subject = str(row[0])
        for metric in metrics:
            position = positions.get(metric)
            if (
                position is not None
                and position < len(row)
                and row[position] is not None
            ):
                value = row[position]
                if isinstance(value, bool):
                    value = str(value)
                elif isinstance(value, (int, float)):
                    value = float(value)
                else:
                    value = str(value)
                yield subject, metric, value, row_number


def _required_number(values: dict[str, Any], field: str) -> float:
    result = _number(_value(values, field, required=True), field=field)
    assert result is not None
    return result


def _package_display(item: str | None, count: float | None) -> str | None:
    """Make a readable display from the workbook's preserved package fields."""
    if item is None or count is None:
        return None
    count_text = str(int(count)) if count.is_integer() else str(count)
    return f"{count_text} × {item}"


def _assert_unique(values: Iterable[str], label: str) -> None:
    entries = list(values)
    if len(entries) != len(set(entries)):
        raise ShopDoctorWorkbookError(f"Workbook contains duplicate {label}")


def _validate(
    items: tuple[WorkbookItem, ...],
    offers: tuple[WorkbookOffer, ...],
    packs: tuple[WorkbookCashPack, ...],
    components: tuple[WorkbookPackComponent, ...],
    relationships: tuple[WorkbookRelationship, ...],
    item_names: set[str],
) -> None:
    _assert_unique([offer.offer_id for offer in offers], "offer IDs")
    _assert_unique([pack.pack_id for pack in packs], "pack IDs")
    unknown = sorted({offer.item_name for offer in offers}.difference(item_names))
    unknown += sorted(
        {component.item_name for component in components}.difference(item_names)
    )
    unknown += sorted(
        {relationship.option_name for relationship in relationships}.difference(
            item_names
        )
    )
    if unknown:
        raise ShopDoctorWorkbookError(
            f"Unknown item catalog references: {', '.join(sorted(set(unknown)))}"
        )
    pack_ids = {pack.pack_id for pack in packs}
    orphaned = sorted(
        {component.pack_id for component in components}.difference(pack_ids)
    )
    if orphaned:
        raise ShopDoctorWorkbookError(f"Orphan pack components: {', '.join(orphaned)}")
    for pack in packs:
        if pack.price_usd is None and pack.assumed_bonus_diamonds is not None:
            raise ShopDoctorWorkbookError(
                f"{pack.pack_id}: bonus diamonds require a cash price"
            )
        if pack.price_usd is not None and pack.assumed_bonus_diamonds != round(
            pack.price_usd * 100
        ):
            raise ShopDoctorWorkbookError(
                f"{pack.pack_id}: assumed bonus Diamonds must round(price × 100)"
            )
    for record in (*offers, *components):
        if record.normalized_quantity <= 0:
            raise ShopDoctorWorkbookError(
                "Normalized quantities must be positive; unknown is not zero"
            )
        if (
            record.normalized_quantity_check is not None
            and abs(record.normalized_quantity - record.normalized_quantity_check)
            > 1e-9
        ):
            raise ShopDoctorWorkbookError(
                "Normalized quantity does not match workbook check"
            )
        if (
            record.actual_item_count is not None
            and record.base_units_per_item is not None
        ):
            if (
                abs(
                    record.normalized_quantity
                    - record.actual_item_count * record.base_units_per_item
                )
                > 1e-9
            ):
                raise ShopDoctorWorkbookError(
                    "Package count × base units does not equal normalized quantity"
                )
        if record.speedup_type and record.actual_package_item != "Speedup":
            raise ShopDoctorWorkbookError(
                "Speedup metadata must retain the canonical Speedup package item"
            )
    for component in components:
        if component.normalization_check and component.normalization_check != "OK":
            raise ShopDoctorWorkbookError(
                f"{component.pack_id}: normalization check is not OK"
            )
    for relation in relationships:
        if relation.contents_rule not in {"CHOOSE ONE", "CONVERTS TO"}:
            raise ShopDoctorWorkbookError(
                f"Unsupported contents rule: {relation.contents_rule}"
            )
        if relation.option_quantity <= 0:
            raise ShopDoctorWorkbookError(
                "Choice/conversion quantities must be positive"
            )
